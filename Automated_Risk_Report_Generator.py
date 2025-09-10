import pandas as pd
from openpyxl import load_workbook

# ========= Step 1: Load Data =========
FILE = r"C:\Users\Sahil Shigwan\Desktop\IDFC_Project\loan_dataset.xlsx"   # your Excel file
df = pd.read_excel(FILE, sheet_name="Loans")

# ========= Step 2: Assumptions =========
pd_map = {"AA":0.002,"A":0.005,"BBB":0.01,"BB":0.02,"B":0.03,"C":0.07,"D":0.2}
lgd = 0.45
ccf_map = {"TermLoan":0, "Revolver":0.75, "WorkingCapital":0.5, "Overdraft":0.2, "SME Loan":0.6}

# ========= Step 3: Compute Metrics =========
df["PD"] = df["Rating"].map(pd_map)
df["CCF"] = df["ProductType"].map(ccf_map)
df["EAD"] = df["Outstanding"] + df["UnusedLimit"]*df["CCF"]
df["EL"] = df["EAD"]*df["PD"]*lgd

# ========= Step 4: Build Report Content =========
kpis = {
    "Total EAD": df["EAD"].sum(),
    "Total EL": df["EL"].sum(),
    "Total Outstanding": df["Outstanding"].sum(),
    "# D-rated names": (df["Rating"]=="D").sum()
}

top10 = df.groupby("Counterparty")[["EAD","EL"]].sum().sort_values("EAD",ascending=False).head(10)
by_product = df.groupby("ProductType")[["EAD","EL"]].sum().sort_values("EAD",ascending=False)
drated = df.loc[df["Rating"]=="D","Counterparty"].drop_duplicates()
limit_breach = df.loc[df["Utilization"]>1,["Counterparty","Utilization"]]

# ========= Step 5: Write Report =========
wb = load_workbook(FILE)

# delete old Report sheet if exists
if "Report" in wb.sheetnames:
    del wb["Report"]
ws = wb.create_sheet("Report")

# KPIs
ws["A1"] = "Daily Credit Snapshot"
row = 3
for k,v in kpis.items():
    ws[f"A{row}"] = k; ws[f"B{row}"] = round(v,2); row+=1

# Top10 Counterparties
ws["A8"] = "Top 10 Counterparties by EAD"
r=9
for idx,rowv in top10.iterrows():
    ws[f"A{r}"] = idx
    ws[f"B{r}"] = round(rowv["EAD"],2)
    ws[f"C{r}"] = round(rowv["EL"],2)
    r+=1

# Exposure by ProductType
ws["E8"] = "Exposure by ProductType"
r=9
for idx,rowv in by_product.iterrows():
    ws[f"E{r}"] = idx
    ws[f"F{r}"] = round(rowv["EAD"],2)
    ws[f"G{r}"] = round(rowv["EL"],2)
    r+=1

# D-rated names beside KPIs
ws["H3"] = "D-rated Names"
r=4
if drated.empty:
    ws["H4"] = "None"
else:
    for name in drated:
        ws[f"H{r}"] = name
        r+=1

# Limit breaches
ws["A22"] = "Limit Breaches (Utilization>1)"
r=23
for idx,rowv in limit_breach.iterrows():
    ws[f"A{r}"] = rowv["Counterparty"]
    ws[f"B{r}"] = round(rowv["Utilization"],2)
    r+=1

wb.save(FILE)
print(f"Report sheet updated successfully in file: {FILE}")
